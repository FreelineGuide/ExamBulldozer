from typing import List, Dict, Any, Union
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
from datetime import datetime
import io
import logging
import os
from config_manager import ConfigManager

class ExcelExporter:
    """Excel导出类"""
    
    def __init__(self):
        """初始化Excel导出器"""
        self.logger = logging.getLogger(__name__)
        self.config = ConfigManager()
        self.header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        self.header_font = Font(bold=True)

    def _format_single_choice(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """格式化单选题数据"""
        formatted = {
            "题目": data["question"],
            "题型": "单选题",
            "答案": data["answer"]
        }
        
        # 添加选项
        for key, value in data["options"].items():
            formatted[f"选项{key}"] = value
        
        # 添加解析（如果有）
        if "analysis" in data:
            formatted["解析"] = data["analysis"]
        
        return formatted
    
    def _format_multiple_choice(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """格式化多选题数据"""
        formatted = {
            "题目": data["question"],
            "题型": "多选题",
            "答案": ",".join(data["answer"])
        }
        
        # 添加选项
        for key, value in data["options"].items():
            formatted[f"选项{key}"] = value
        
        # 添加解析（如果有）
        if "analysis" in data:
            formatted["解析"] = data["analysis"]
        
        return formatted
    
    def _format_true_false(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """格式化判断题数据"""
        formatted = {
            "题目": data["question"],
            "题型": "判断题",
            "答案": "对" if data["answer"] else "错"
        }
        
        # 添加解析（如果有）
        if "analysis" in data:
            formatted["解析"] = data["analysis"]
        
        return formatted
    
    def _format_custom(self, data: Dict[str, Any], question_type: str) -> Dict[str, Any]:
        """格式化自定义题型数据"""
        formatted = {
            "题目": data["question"],
            "题型": question_type
        }
        
        # 添加选项（如果有）
        if "options" in data:
            if isinstance(data["options"], dict):
                for key, value in data["options"].items():
                    formatted[f"选项{key}"] = value
            elif isinstance(data["options"], list):
                for i, value in enumerate(data["options"]):
                    formatted[f"选项{i+1}"] = value
        
        # 添加答案
        if "answer" in data:
            if isinstance(data["answer"], list):
                formatted["答案"] = ",".join(map(str, data["answer"]))
            else:
                formatted["答案"] = str(data["answer"])
        
        # 添加解析（如果有）
        if "analysis" in data:
            formatted["解析"] = data["analysis"]
        
        return formatted
    
    def _get_export_filename(self, question_type: str) -> str:
        """生成导出文件名"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{question_type}_{timestamp}.xlsx"

    def export_to_excel(self, data: Union[Dict[str, Any], List[Dict[str, Any]]], question_type: str) -> tuple[bytes, str]:
        """导出数据到Excel，返回字节流和文件名"""
        try:
            self.logger.info(f"开始导出{question_type}类型的试题数据")
            
            # 确保数据是列表格式
            if isinstance(data, dict):
                data = [data]
            
            # 格式化数据
            formatted_data = []
            for item in data:
                try:
                    if question_type == "single_choice":
                        formatted = self._format_single_choice(item)
                    elif question_type == "multiple_choice":
                        formatted = self._format_multiple_choice(item)
                    elif question_type == "true_false":
                        formatted = self._format_true_false(item)
                    else:
                        formatted = self._format_custom(item, question_type)
                    formatted_data.append(formatted)
                except Exception as e:
                    self.logger.error(f"格式化数据项失败: {str(e)}")
                    raise Exception(f"格式化数据失败：{str(e)}")
            
            if not formatted_data:
                raise Exception("没有可导出的数据")
            
            # 创建DataFrame
            df = pd.DataFrame(formatted_data)
            
            # 生成文件名
            filename = self._get_export_filename(question_type)
            
            # 创建字节流
            excel_buffer = io.BytesIO()
            
            # 导出到字节流
            self.logger.info("正在生成Excel文件")
            df.to_excel(excel_buffer, index=False, engine='openpyxl')
            
            # 获取字节数据
            excel_data = excel_buffer.getvalue()
            
            self.logger.info("导出成功")
            return excel_data, filename
            
        except Exception as e:
            self.logger.error(f"导出Excel失败：{str(e)}")
            raise Exception(f"导出Excel失败：{str(e)}")

    def get_default_filename(self) -> str:
        """获取默认的Excel文件名"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"试题数据_{timestamp}.xlsx"

    def set_export_path(self, path: str) -> None:
        """设置导出路径"""
        self.config.set_export_path(path)
        self.export_path = path 